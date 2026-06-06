
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd

dados = pd.read_csv("dados.csv")

print("\n=== STATUS EXISTENTES ===")
print(dados["status"].unique())

# filtros
sem_ticket = dados[dados["ticket"].isna()]
abertos = dados[dados["status"] == "Aberto"]
concluidos = dados[dados["status"] == "Concluído"]
pendentes = dados[dados["status"] == "Pendente"]
em_andamento = dados[dados["status"] == "Em andamento"]

resumo = pd.DataFrame({
    "Indicador": [
        "Total de empresas",
        "Sem ticket",
        "Abertos",
        "Concluídos",
        "Pendentes",
        "Em andamento"
    ],
    "Quantidade": [
        len(dados),
        len(sem_ticket),
        len(abertos),
        len(concluidos),
        len(pendentes),
        len(em_andamento)
    ]
})

# resumo
print("\n=== RESUMO ===")
print(f"Total de empresas: {len(dados)}")
print(f"Sem ticket: {len(sem_ticket)}")
print(f"Abertos: {len(abertos)}")
print(f"Concluídos: {len(concluidos)}")
print(f"Pendentes: {len(pendentes)}")
print(f"Em andamento: {len(em_andamento)}")

print("\n=== EMPRESAS QUE PRECISAM DE TICKET ===")

for _, linha in sem_ticket.iterrows():
    print(
        f"{linha['empresa']} precisa de abertura de ticket - problema: {linha['problema']}"
    )

# gerar excel
with pd.ExcelWriter("relatorio.xlsx") as writer:
    dados.to_excel(writer, sheet_name="Dados", index=False)
    resumo.to_excel(writer, sheet_name="Resumo", index=False)
print("\nRelatório Excel gerado com sucesso!")


# abrir o arquivo excel
wb = load_workbook("relatorio.xlsx")
# CORES DOS STATUS
verde = PatternFill("solid", fgColor="C6EFCE")
amarelo = PatternFill("solid", fgColor="FFF2CC")
vermelho = PatternFill("solid", fgColor="F4CCCC")
azul = PatternFill("solid", fgColor="D9EAF7")
laranja = PatternFill("solid", fgColor="FCE5CD")

# formata todas as abas
for aba in wb.worksheets:

    # cabeçalho em negrito e centralizado
    for celula in aba[1]:
        celula.font = Font(bold=True)
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        celula.fill = PatternFill(
            fill_type="solid",
            fgColor="4472c4"
        )

        # Criar borda fina
        borda = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

 # Aplicar borda em todas as células
    for linha in aba.iter_rows():
        for celula in linha:
            celula.border = borda

            # Todas as células centralizadas
    for linha in aba.iter_rows():
        for celula in linha:
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            # COLORIR STATUS
    for linha in aba.iter_rows(min_row=2):

        status = linha[-1].value

        if status == "Concluído":
            linha[-1].fill = verde

        elif status == "Aberto":
            linha[-1].fill = amarelo

        elif status == "Sem ticket":
            linha[-1].fill = vermelho

        elif status == "Em andamento":
            linha[-1].fill = azul

        elif status == "Pendente":
            linha[-1].fill = laranja

    # ajustar largura das colunas
    for coluna in aba.columns:
        tamanho_max = 0
        letra_coluna = coluna[0].column_letter

        for celula in coluna:
            try:
                if len(str(celula.value)) > tamanho_max:
                    tamanho_max = len(str(celula.value))

            except:
                pass

            aba.column_dimensions[letra_coluna].width = tamanho_max + 3

            # Congelar primeira linha
            aba.freeze_panes = "A2"

            # salvar alterações
            wb.save("relatorio.xlsx")
